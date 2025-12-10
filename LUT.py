import openpyxl
import math
import numpy as np

def convert_to_q4_10(float_value):
    """
    將一個浮點數轉換為 Q4.10 格式的定點整數。
    """
    if not isinstance(float_value, (int, float)):
        return "Invalid Input"

    integer_bits = 4
    fractional_bits = 10
    total_bits = integer_bits + fractional_bits
    min_val = - (2**(total_bits - 1))
    max_val = (2**(total_bits - 1)) - 1
    scaling_factor = 2**fractional_bits

    fixed_point_integer = round(float_value * scaling_factor)
    
    if fixed_point_integer > max_val:
        # return f"Overflow ({max_val})"
        return max_val
    if fixed_point_integer < min_val:
        # return f"Underflow ({min_val})"
        return min_val
        
    return fixed_point_integer

# 建立一個新的 Excel 工作簿
workbook = openpyxl.Workbook()

# 選擇預設的工作表
sheet = workbook.active
sheet.title = "E_Log_LUT"

print("開始生成表格...")

# --- 步驟 1: 寫入欄位標題 ---
# sheet["A1"] = "Lm(int16)"
sheet["A1"] = "E(int8)"
# 在標題中加入公式，方便日後理解
# sheet["B1"] = "log(Lm)(q4.10)"
sheet["B1"] = "(E-144)log2(q4.10)"

# --- 步驟 2: 進行迴圈與計算 ---

# 迴圈範圍是從 0 到 2^20 - 1
# 2**20 代表 2 的 20 次方
LOG10_2 = np.log10(2)
total_rows = 2**8
for x in range(total_rows):
  # 提取高位元部分 (bits 19-4)
  # x >> 4 相當於將 x 的二進位右移 4 位
  # val_high_bits = x >> 4

  # 提取低位元部分 (bits 3-0)
  # x & 15 (二進位 1111) 只保留最右邊 4 個位元
  # val_low_bits = x & 15

  # --- 進行計算 ---
  
  # 檢查高位元部分是否為 0，以避免 log(0) 的數學錯誤
  # 當 x 的值在 0 到 15 之間時，高位元部分會是 0
  # if x == 0:
  #   result = "Error: log(0)"
  # else:
  #   try:
  #     計算 2 的次方項
  #     power_of_two = math.pow(2, val_low_bits - 16)
  #     計算 log 的參數
  #     argument = val_high_bits * power_of_two
  #     計算自然對數 (log base e)
  #     result = math.log(argument)
  #     如果您需要以 10 為底的對數，請將上一行改成下面這行：
  result = (x - 144) * LOG10_2  
      # result = math.log10(x)
    # except ValueError:
    #   # 處理其他可能的數學錯誤
    #   result = "Error: Math Domain"

  q4_10_result = convert_to_q4_10(result)
  # --- 將值寫入 Excel ---
  # Excel 的列是從 1 開始，且第一列是標題，所以要 +2
  current_row = x + 2
  sheet.cell(row=current_row, column=1).value = x
  sheet.cell(row=current_row, column=2).value = q4_10_result

  # 每處理 100,000 列後印出進度，讓你知道程式正在執行
  if (x + 1) % 100000 == 0:
    print(f"已處理 {x + 1} / {total_rows} 列...")


# --- 步驟 3: 儲存檔案 ---
print("計算完成，正在儲存檔案... 這可能會需要一些時間。")
workbook.save("E_log_LUT.xlsx")

print(f"成功生成 log_calculation_1D.xlsx 檔案！總共包含 {total_rows} 筆資料。")
